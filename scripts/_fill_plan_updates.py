"""One-shot: populate the Updates column in 'DCC AI Agent Plan.xlsx' Sheet4
for every row whose Status is not 'Planned'. Reads the current state of the
codebase and writes a milestone summary per row.
"""
import sys
sys.stdout.reconfigure(encoding="utf-8")

import openpyxl
from openpyxl.styles import Alignment

WB_PATH = "DCC AI Agent Plan.xlsx"

UPDATES_COL = 9  # Column I

UPDATES = {
    5: (
        "The bot understands what the patient is complaining about — in Arabic or English, including Saudi and Egyptian dialects — and routes them to the right medical specialty. "
        "It recognises symptoms, body parts, and direct specialty names, and it handles typos, slang, and informal phrasing. "
        "More than 75 specialties are supported. Currently being validated across a wide set of realistic patient scenarios before full rollout."
    ),
    6: (
        "The bot pulls live doctor lists and their availability directly from the hospital system, so every recommendation reflects the real schedule at that moment. "
        "Patients can choose a doctor by number, by full name, or by a partial / misspelled name in either language, and the bot will still find the right match. "
        "This integration is currently read-only — the bot reads availability but does not yet write the booking back into the hospital system."
    ),
    7: (
        "Patients can describe the time they want in natural language — 'tonight', 'after 7', 'morning', a specific hour, or just 'the earliest one'. "
        "When the exact time isn't available, the bot suggests the closest match and waits for the patient to confirm. "
        "Once the time is agreed, the bot collects the phone number and insurance details and produces a full booking confirmation message. "
        "The confirmation is generated and shown to the patient; writing it back into the hospital system is the next step."
    ),
    8: (
        "Walk-in (cash) prices are shown per doctor when the patient asks. "
        "The bot handles the three common cases: cash patients get a quote, insured patients are told a representative will call back with the price, and patients who haven't said yet are asked to clarify. "
        "Tiered pricing, offers, promotional rates, and insurance-specific pricing are not yet in scope."
    ),
    9: (
        "The bot is deployed on the internal company network and is accessible to the team for hands-on testing. "
        "It runs against the real hospital data (live doctor schedules, real availability, real walk-in prices). "
        "A stakeholder view is built into the interface so non-technical team members can demo the bot and inspect what it understood from each message. "
        "No external / patient-facing access yet — this is a controlled internal rollout."
    ),
    10: (
        "A library of scripted patient scenarios runs end-to-end against the bot on demand. "
        "It covers symptom-based bookings, specialty-based bookings, booking by doctor name, and several tricky edge cases (price asked mid-booking, combined answers, etc.). "
        "Each run produces a full transcript, a summary of what changed at every turn, and timing / cost metrics — making it easy to spot regressions and measure improvements between versions."
    ),
    11: (
        "The scenario runs are being used to find where the bot breaks or behaves poorly. "
        "The latest baseline round surfaced five recurring issue categories — all five have been addressed this week with targeted fixes. "
        "Known remaining gaps: some doctors don't have a walk-in price on file (patient gets a callback instead of a quote), and insurance handling today is accept / reject by company name only. "
        "Larger pending items (inquiry handling, cancellation, rescheduling, writing bookings back into the hospital system, WhatsApp channel) are tracked separately."
    ),
    12: (
        "The instructions given to the AI have been refined repeatedly based on real scenario results. "
        "The bot now replies in the same language as the patient, keeps messages short and on-brand, matches the tone of a real Andalusia agent, and stays on track during the booking flow instead of drifting into unrelated topics. "
        "Cost per conversation is kept low by only showing the AI the most recent part of the chat."
    ),
    13: (
        "Core reference data is organised and ready: the full list of specialties in both languages, the mapping from symptoms to specialties, the list of accepted insurance companies (with Arabic spellings), and every patient-facing message in Arabic and English. "
        "Still to be added: a searchable FAQ / policy repository, descriptions of services and procedures, and working-hours / clinic-info content. "
        "These are the foundations for the upcoming Inquiry module."
    ),
    14: (
        "Walk-in cash price is structured per doctor and available to the bot in real time. "
        "Other pricing dimensions are not yet structured: service-vs-consultation tiers, offer / campaign pricing, validity windows, eligibility rules, and insurance-specific rates. "
        "The data pipeline is ready to accept this richer pricing data once the business side provides it."
    ),
    17: (
        "The bot recognises 25+ accepted insurance companies and handles misspellings and Arabic spellings of each. "
        "Once a patient names their insurer, the bot routes them to a callback flow instead of quoting a cash price. "
        "Deeper eligibility checks — policy validation, per-service coverage rules, real-time eligibility API with the insurer — are not yet in place. Today the decision is accept / reject by company name only."
    ),
    18: (
        "Not yet implemented. "
        "The groundwork is in place (the bot can already look up per-doctor pricing and hospital data in real time), but the offers / campaigns logic itself — discount rules, promotional pricing, eligibility by patient or service, validity windows — has not been built yet. "
        "This milestone is at the design stage."
    ),
}


def main() -> None:
    wb = openpyxl.load_workbook(WB_PATH)
    ws = wb["Sheet4"]

    for row_num, text in UPDATES.items():
        cell = ws.cell(row=row_num, column=UPDATES_COL)
        cell.value = text
        cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
        ws.row_dimensions[row_num].height = None

    wb.save(WB_PATH)
    print(f"Wrote updates to {len(UPDATES)} rows")

    wb2 = openpyxl.load_workbook(WB_PATH, data_only=True)
    ws2 = wb2["Sheet4"]
    for row_num in sorted(UPDATES.keys()):
        v = ws2.cell(row=row_num, column=UPDATES_COL).value
        milestone = ws2.cell(row=row_num, column=3).value
        status = (ws2.cell(row=row_num, column=5).value or "").strip().split("\n")[0]
        first_line = (v or "").split("\n")[0][:65]
        print(f"  Row {row_num:2d} [{status:38.38}] {milestone:42.42} -> {first_line}")


if __name__ == "__main__":
    main()
